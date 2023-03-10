#!/usr/bin/env python
# -*- coding=utf-8 -*-


import base64
import sys
import os
import yaml
import json
import re
import requests
import time
from collections import defaultdict
from xml.etree.ElementTree import parse
import logging
from lib import package
from txdpy import get_Bletter, get_Sletter
import xlwt
import xlrd
from openpyxl import load_workbook
from xlutils.copy import copy
import copy
from lxml import html


srcOepkgsNum = 0
allFileNum = 0
allFileList = []  # 存放 当前路径 以及当前路径的子路径 下的所有文件
sigInfolist = []
allYamlList = []
filter_char = []
allYamldata = []
src_code_up = []
src_code_is = []
allYamldata_tag = []
yaml_error = []
Inyaml = []
dict_oepkgs = {}
d_dict = defaultdict(list)
d_oepkg = defaultdict(list)
dict_list = defaultdict(list)


headers = {"Content-Type": "application/json;charset=UTF-8"}
real_path = os.path.dirname(os.path.realpath(__file__)) + "/"
closed_header = "curl -X PATCH --header 'Content-Type: application/json;charset=UTF-8'"
rq_header = "curl -X POST --header 'Content-Type: application/json;charset=UTF-8'"
api_token = "c4a7f2254bd58885a9c6fa80cbd0b7dc"
openeuler_version = ["openEuler-20.03-LTS", "openEuler-20.03-LTS-SP1", "openEuler-20.03-LTS-SP2", "openEuler-20.03-LTS-SP3", "openEuler-20.09", "openEuler-21.03", "openEuler-21.09", "openEuler-22.03-LTS", "openEuler-22.03-LTS-SP1", "openEuler-22.09"]
oepkgs_version = ["openeuler-20.03-LTS-SP1", "openeuler-20.03-LTS-SP2", "openeuler-20.03-LTS-SP3", "openeuler-22.03-LTS", "openeuler-22.03-LTS-SP1"]


def lib_data(version):
    if version is None:
        sys.exit()
    rpm_pkg_route = "/srv/rpm/testing/{0}".format(version)
    # 取rpm包总数和rpm文件绝对路径
    package.getAllFilesInPath(rpm_pkg_route)
    logging.info("当前路径下的总文件数 =", allFileNum)
    for rpm_path in allFileList:
        rpm_file = package.shell_cmd("Name", rpm_path)  # 获取rpm信息
        d[rpm_file].append(rpm_path)
    logging.info('--------rpm file-----------')
    # 取一个版本的的所有包的信息以字典形式存入json文件
    with open("sp3_yaml.json", "w") as fw:
        fw.write(json.dumps(d))


def pr_bat():
    num = 0
    # 创建pr
    os.system("git clone 'https://gitee.com/zhang-yn/oepkgs-management.git';")
    package.getAllFilesInPath_1("./oepkgs-management_10/sig")
    package.getAllFilesInPath_1("./oepkgs-management/sig")
    for i, item in enumerate(allYamldata):
        if len(get_Bletter(item.split("/")[-1][:-5])) != 0:
            if item.split("/")[-1][:-5].lower() + ".yaml" in Inyaml:
                continue
        dest_path = real_path + "oepkgs-management" + "/" + "/".join(item.split("/")[-5:-1]) + "/"
        if item.split("/")[-1] not in Inyaml and len(item.split("/")[-1][:-5]) > 1 and len(
                item.split("/")[-1].split(".")) == 2:
            num = num + 1
            if not os.path.exists(dest_path):
                os.system("mkdir -p {0};cp -rf {1} {0}".format(dest_path, item))
            else:
                os.system("cp -rf {1} {0}".format(dest_path, item))
            if num >= 100:
                os.system("cd {0};git add .;git commit -m '自动化仓库创建';git push".format("oepkgs-management"))
                package.creat_pr()
                time.sleep(500)
                num = 0
    logging.info("------- creat end -------")


def upload_rpmcode(data):
    for yaml_f in data:
        module_name = data[yaml_f]
        rpm_dict = {}
        version_set = set()
        for rpm_route in module_name:
            rpm_version = package.shell_cmd("Version", rpm_route)
            version_set.add(rpm_version)
            rpm_dict[rpm_version] = rpm_route
        version_list = list(version_set)
        version_list.sort()
        os.system("curl -X DELETE --header 'Content-Type: application/json;charset=UTF-8' 'https://gitee.com/api/v5/repos/src-oepkgs/{}/branches/{}/setting?access_token={}'".format(yaml_f, sys.argv[1], api_token))
        for rpm_version in version_list:
            rpm_route = rpm_dict[rpm_version]
            package.push_pkg(yaml_f, rpm_route, rpm_version)
        os.system("curl -X PUT --header 'Content-Type: application/json;charset=UTF-8' 'https://gitee.com/api/v5/repos/src-oepkgs/{}/branches/{}/protection' -d '{{\"access_token\":\"{}\"}}'".format(yaml_f, sys.argv[1], api_token))
        src_code_up.append(yaml_f)


def yaml_part(data):
    tag_num = 0
    add_yaml = 0
    for yaml_data in data:
        tag_num = tag_num + 1
        logging.info("------ branch {} 已添加 -----".format(tag_num))
        yaml_file = package.yamlName(yaml_data)
        if yaml_file not in allYamlList:
            src_code_is.append(yaml_file)
            continue
        os.system("git clone 'https://gitee.com/src-oepkgs/{0}.git';".format(yaml_file))
        if not os.path.exists(real_path + yaml_file):
            yaml_file = yaml_file.lower()
            os.system("git clone 'https://gitee.com/src-oepkgs/{0}.git';".format(yaml_file))
            if not os.path.exists(real_path + yaml_file):
                allYamldata.append(yaml_file)
                logging.info("------ allYamldata {} 已添加 -----".format(yaml_file))
                continue
        os.chdir(os.getcwd() + "/" + yaml_file)
        commit_id = os.popen("git tag").read().strip()
        repo_branch = os.popen("git branch").read().strip()
        if repo_branch == "":
            logging.info("----- {} branch 不存在 -----".format(yaml_file))
            os.chdir(os.path.pardir)
            os.system("rm -rf {0}".format(yaml_file))
            continue
        if commit_id == "":
            add_yaml = package.judge_commitId(yaml_file, add_yaml, yaml_data, data)
        else:
            tag_list = commit_id.split("\n")
            if "20.03-LTS-SP3" in [i[:13] for i in tag_list]:
                logging.info("----- {} tag 已存在 -----".format(yaml_file))
                allYamldata_tag.append(yaml_file)
                os.chdir(os.path.pardir)
                os.system("rm -rf {0}".format(yaml_file))
                continue
            else:
                d_oepkg[yaml_file] = d[yaml_data]
                os.chdir(os.path.pardir)
                os.system("rm -rf {0}".format(yaml_file))
                add_yaml = add_yaml + 1
                logging.info("------ {0} branch {1} 已添加 -----".format(yaml_file, add_yaml))


def rw_xsl():
    xl = xlwt.Workbook(encoding='UTF-8')
    sheet = xl.add_sheet('docker', cell_overwrite_ok=True)
    row = 0
    cloum = 0

    # 取输入文件的每一行进行excel插入
    with open("输入文件.txt", "r") as f:
        for k, i in enumerate(f.readlines()):
            a = os.popen("grep FROM {}".format(i)).read()
            data = i.split("/")
            for j in data:
                sheet.write(row, cloum, j)
                cloum = cloum + 1
            sheet.write(row, cloum, a)
            cloum = cloum + 1
            with open("docker_link.txt", "r") as f1:
                sheet.write(row, cloum, f1.readlines()[k])
            row = row + 1
            cloum = 0
    #写入docker
    xl.save("docker.xls")


def count():
    book = xlrd.open_workbook("oepkgs.xlsx")
    # 获取第一张工作表
    sh = book.sheet_by_index(3)
    col_value = sh.col_values(0)
    del col_value[0]
    del col_value[1]
    del col_value[2]
    xls = xlrd.open_workbook("oepkgs.xlsx")
    xls_file = copy(xls)
    sheet = xls_file.get_sheet(3)
    with open("test.json", "r", ) as f:
        openeuler_data = json.loads(f.read())
    with open("oepkgs.json", "r", ) as f1:
        oepkgs_data = json.loads(f1.read())
    oepkgs_list = []
    oepkgs_list1 = []
    a_list = []
    c = []
    for value_name in col_value:
        a_list.append(value_name.lower())
    name_list = ["oepkgs_data", "openeuler_data"]
    for number in name_list:
        for key in oepkgs_data.keys():
            for j in oepkgs_data[key].keys():
                oepkgs_list.append(j)
    for oepkg_value in oepkgs_list:
        oepkgs_list1.append(oepkg_value.lower())
    for value_a in a_list:
        if value_a not in oepkgs_list1:
            c.append(i)


def excel_insert():
    suse_name = {}
    book = xlrd.open_workbook("zyn1.xlsx")
    # 获取第一张工作表
    sh = book.sheet_by_index(0)
    col_value = sh.col_values(0)
    del col_value[0]
    # 读取
    wbook = openpyxl.load_workbook("zyn1.xlsx")
    sheet = wbook['包清单']
    a_list = []
    # num为要写入的excel文件的行数+1
    num_index = 124756
    for key in suse_name:
        if key.lower() not in col_value:
            sheet["A{}".format(num_index)] = key.lower()
            sheet["B{}".format(num_index)] = "SUSE"
            sheet["C{}".format(num_index)] = "FALSE"
            sheet["D{}".format(num_index)] = "FALSE"
            sheet["E{}".format(num_index)] = "FALSE"
            sheet["F{}".format(num_index)] = "FALSE"
            sheet["G{}".format(num_index)] = "FALSE"
            sheet["H{}".format(num_index)] = "FALSE"
            sheet["I{}".format(num_index)] = "FALSE"
            sheet["G{}".format(num_index)] = "FALSE"
            sheet["K{}".format(num_index)] = "FALSE"
            sheet["L{}".format(num_index)] = "FALSE"
            sheet["M{}".format(num_index)] = "FALSE"
            sheet["N{}".format(num_index)] = "FALSE"
            sheet["O{}".format(num_index)] = "FALSE"
            sheet["P{}".format(num_index)] = "FALSE"
            sheet["Q{}".format(num_index)] = "FALSE"
            sheet["R{}".format(num_index)] = "FALSE"
            sheet["S{}".format(num_index)] = "FALSE"
            sheet["T{}".format(num_index)] = "TRUE"
            num_index = num_index + 1
        else:
            sheet["T{}".format(col_value.index(key.lower()) + 2)] = "TRUE"
    wbook.save("zyn2.xlsx")


def group():
    # 创建pr
    os.system("git clone 'https://gitee.com/zhang-yn/oepkgs-management.git';")
    package.getAllFilesInPath_1("./oepkgs-management/sig")
    wbk = xlwt.Workbook()
    ws = wbk.add_sheet('1 sheet')
    line = 0
    for vaule_i in Inyaml:
        package.read_yaml(vaule_i, ws, line)
        line = line + 1
    logging.info("------ test -----")
    wbk.save('1.xls')


def oepkgs_apply_pro():
    oepkgs_list = []
    col_value1, sheet1, xls_file1 = package.open_file()
    for value_element in col_value1:
        oepkgs_list.append(value_element.lower())
    with open("test.json", "r", ) as f:
        openeuler_data = json.loads(f.read())
    with open("oepkgs.json", "r", ) as f1:
        oepkgs_data = json.loads(f1.read())
    col = 2
    col = package.write(col, oepkgs_data, sheet1)
    package.write(col, openeuler_data, sheet1)
    xls_file1.save("euler_pkg.xls")


if __name__ == "__main__":
    import argparse
    par = argparse.ArgumentParser()
    par.add_argument("-s", "--script", help="script name", required=True)
    par.add_argument("-v", "--version", help="rpm package version", required=False)
    args = par.parse_args()
    if args.script == "lib_data":
        lib_data(args.version)
    elif args.script == "pr_bat":
        pr_bat()
    elif args.script == "upload_rpmcpde":
        with open("yaml_sp3.json", "r") as ff:
            file_data = json.load(ff)
        upload_rpmcode(file_data)
    elif args.script == "yaml_part":
        package.getAllFilesInPath("./oepkgs-management/sig")
        # sp3_yaml.json是lib_data.py脚本生成的一个版本包的信息
        with open("sp3_yaml.json", "r") as fj:
            d = json.load(fj)
        yaml_part(d)
        with open("yaml_sp3.json", "w") as fb:
            fb.write(json.dumps(d_oepkg))
        logging.info("--------- yaml list -----------")
    elif args.script == "re_xls":
        rw_xsl()
    elif args.script == "count":
        count()
    elif args.script == "excel_insert":
        excel_insert()
    elif args.script == "group":
        # 创建pr
        group()
    elif args.script == "oepkgs_apply":
        package.openeuler(openeuler_version)
        package.oepkgs(oepkgs_version)
    elif args.script == "oepkgs_apply_pro":
        oepkgs_apply_pro()
    elif args.script == "creat_file":
        rpm_pkg_path = "/srv/rpm/pub/"
        package.getAllFilesInPath(rpm_pkg_path)
        for rpm_path in allFileList:
            rpm_file = package.shell_cmd("Name", rpm_path)  # 获取rpm信息
            d_dict[rpm_file].append(rpm_path)
        os.system("git clone 'https://gitee.com/zhang-yn/oepkgs-management.git';")
        # 获取src-oepkgs上已经存在的库，通过yaml文件获取
        package.getAllFilesInPath("./oepkgs-management/sig")
        # 判断取到的rpm文件是否在舱内已经存在，进行过滤)
        d_list = copy.deepcopy(d_dict)
        for d_list_name in d_list:
            if d_list_name in allYamlList or str.lower(d_list_name) in allYamlList:
                d_dict.pop(d_list_name)
        # 遍历字典进行yaml创建
        package.source_code("module2.xml")
        for yaml_modify in d_dict:
            yaml_name = package.yamlName(yaml_modify)
            package.data_box(yaml_modify,yaml_name)
        logging.info("------- 剩余 ------")
        for oepkg_keys in d_oepkg.keys():
            if not os.path.exists(real_path + "/oepkgs-management/sig/{}/sig-info.yaml".format(oepkg_keys)):
                package.yaml_isexist(oepkg_keys)
            else:
                package.yaml_not_exist(oepkg_keys)