#!/usr/bin/env python3
# coding=utf-8
import urllib3
from queue import Queue
import threading
import re
import html
import openpyxl

'''
整体的思路：
1、构造任务队列pageQueue ，存放所有要爬取的页面url
2、用多线程爬虫将抓取的页面内容存放到data_queue中
3、用多线程程序对data_queue中的页面内容进行解析，分别提取id, product, os, company, category，
'''

class Crawl_thread(threading.Thread):
    '''
    抓取线程类，注意需要继承线程类Thread
    '''
    def __init__(self,thread_id,queue):
        threading.Thread.__init__(self)
        self.thread_id = thread_id
        self.queue = queue # 任务队列

    def run(self):
        '''
        线程在调用过程中就会调用对应的run方法
        :return:
        '''
        print('启动线程：',self.thread_id)
        self.crawl_spider()
        print('退出了该线程：',self.thread_id)

    def crawl_spider(self):
        while True:
            if self.queue.empty(): #如果队列为空，则跳出
                break
            else:
                page = self.queue.get()
                print('当前工作的线程为：',self.thread_id," 正在采集：",page)
                try:
                    http = urllib3.PoolManager(10)
                    url = "https://www.suse.com/nbswebapp/yesBulletin.jsp?bulletinNumber=" + str(page)
                    page_data = http.request('GET', url)
                    status_code = page_data.status
                    if status_code == 200:
                        page_html = page_data.data.decode('utf-8')
                        data_queue.put((page_html, page))
                except Exception as e:
                    print('采集线程错误',e)

class Parser_thread(threading.Thread):
    '''
    解析网页的类，就是对采集结果进行解析，也是多线程方式进行解析
    '''
    def __init__(self,thread_id,queue):
        threading.Thread.__init__(self)
        self.thread_id = thread_id
        self.queue = queue

    def run(self):
        print('启动线程：', self.thread_id)
        while not flag:
            try:
                item = self.queue.get(False) # get参数为false时队列为空，会抛出异常
                if not item:
                    pass
                self.parse_data(item)
            except Exception as e:
                pass
        print('退出了该线程：', self.thread_id)
    def parse_data(self,item):
        '''
        解析网页内容的函数
        :param item:
        :return:
        '''
        try:
            print("开始解析：", self.thread_id)
            page = item[0]
            id = item[1]
            product = ""
            category = ""
            company = ""
            os = ""
            regex = re.compile('<TITLE>(.*?)</TITLE>', re.S)
            result = regex.findall(page)
            regex_category = re.compile('<FONT COLOR="439539"><B>(.*?)</B></FONT>', re.S)
            result_category = regex_category.findall(page)
            regex_company = re.compile('For more information regarding the specific test configuration, please contact:<br>(.*?)<br>(.*?)<FONT SIZE="2"><B>(.*?)</B>', re.S)
            result_company = regex_company.findall(page)
            regex_os = re.compile('<b>Operating Systems:</b>(.*?)<br>(.*?)<table style="font-size: 80%;" border="0" cellpadding="0" cellspacing="0" width="100%">(.*?)<tr>(.*?)<td valign="top">(.*?)</td>', re.S)
            result_os = regex_os.findall(page)

            for item in result:
                product += html.unescape(item.strip()).strip()

            for item in result_category:
                category += html.unescape(item.strip()).strip()

            for item in result_company:
                company += html.unescape(item[-1].strip()).strip()

            for item in result_os:
                os += html.unescape(item[-1].strip()).strip()

            response = {
                'id': id,
                'product': product,
                'category': category,
                'company': company,
                'os': os
            }

            all_data.append(response)
        except Exception as e:
            print('parse: ',e)


data_queue = Queue(50)
all_data = []
flag = False
def main():
    pageQueue = Queue(100) # 任务队列，存放网页的队列
    for page in range(101400, 101500):
        pageQueue.put(page) # 构造任务队列
    # 初始化采集线程
    crawl_threads = []
    crawl_name_list = ['crawl_' + str(i) for i in range(1, 6)] # 总共构造5个爬虫线程
    for thread_id in crawl_name_list:
        thread = Crawl_thread(thread_id, pageQueue) # 启动爬虫线程
        thread.start() # 启动线程
        crawl_threads.append(thread)

    # 初始化解析线程
    parse_thread = []
    parser_name_list = ['parse_' + str(i) for i in range(1, 6)]
    for thread_id in parser_name_list:
        thread = Parser_thread(thread_id,data_queue)
        thread.start() # 启动线程
        parse_thread.append(thread)

    # 等待队列情况，先进行网页的抓取
    while not pageQueue.empty():
        pass # 不为空，则继续阻塞

    # 等待所有线程结束
    for t in crawl_threads:
        t.join()
    # 等待队列情况，对采集的页面队列中的页面进行解析，等待所有页面解析完成
    while not data_queue.empty():
        print("队列里面没有数据")
        pass
    # 通知线程退出
    global flag
    flag = True
    for t in parse_thread:
        t.join() # 等待所有线程执行到此处再继续往下执行

    wb = openpyxl.Workbook()
    sheet = wb.active

    for col, head in enumerate(all_data[0].keys()):
        sheet.cell(row=1, column=col+1, value=head)

    for row, sheet_row in enumerate(all_data):
        for col, row_key in enumerate(sheet_row.keys()):
            sheet.cell(row=row+2, column=col+1, value=sheet_row[row_key])
    wb.save(filename="suse_compatiable_list.xlsx")

if __name__ == '__main__':
    main()
