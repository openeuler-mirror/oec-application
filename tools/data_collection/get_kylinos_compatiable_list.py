#!/usr/bin/env python3
# coding=utf-8
import urllib3
import json
import openpyxl
from queue import Queue
import threading

'''
整体的思路：
1、构造任务队列pageQueue ，存放所有要爬取的页面url
2、用多线程爬虫，然后将抓取的页面内容存放到all_data中
3、然后存放到的xlsx文件中
'''

class Crawl_thread(threading.Thread):
    '''
    抓取线程类，注意需要继承线程类Thread
    '''
    def __init__(self,thread_id,queue):
        threading.Thread.__init__(self) # 需要对父类的构造函数进行初始化
        self.thread_id = thread_id
        self.queue = queue # 任务队列

    def run(self):
        '''
        线程在调用过程中就会调用对应的run方法
        :return:
        '''
        print('启动线程：',self.thread_id)
        self.crawl_spider()
        print('退出了该线程：',self.thread_id)

    def crawl_spider(self):
        while True:
            if self.queue.empty(): #如果队列为空，则跳出
                break
            else:
                page = self.queue.get()
                print("第 %d 页" %(page))
                print('当前工作的线程为：',self.thread_id," 正在采集：",page)
                try:
                    http = urllib3.PoolManager(10)
                    url = "https://eco.kylinos.cn/home/compatible/index.html?system_class=1" \
                          "&system_id=&small_version_id=&is_plan=0&page=" + str(page) + "&limit=20"
                    req_data = json.loads(http.request('GET', url).data.decode('utf-8'))["data"]
                    all_data.extend(req_data)
                    print("数据队列长度：",len(all_data),end="\n")
                except Exception as e:
                    print('采集线程错误',e)

all_data = []
def main():
    http = urllib3.PoolManager(10)
    req = http.request('GET', 'https://eco.kylinos.cn/home/compatible/index.html?'
                              'system_class=1&system_id=&small_version_id=&is_plan=0&page=1&limit=20')

    encoded_data = json.loads(req.data.decode('utf-8'))
    count = encoded_data["count"]
    page_num = (count // 20) + 1
    pageQueue = Queue(count) # 任务队列，存放网页的队列
    for page in range(1, page_num+1):
        pageQueue.put(page) # 构造任务队列
    # 初始化采集线程
    crawl_threads = []
    crawl_name_list = ['crawl_' + str(i) for i in range(1, 6)] # 总共构造5个爬虫线程
    for thread_id in crawl_name_list:
        thread = Crawl_thread(thread_id, pageQueue) # 启动爬虫线程
        thread.start() # 启动线程
        crawl_threads.append(thread)

    # 等待队列情况，先进行网页的抓取
    while not pageQueue.empty():
        # 不为空，则继续阻塞
        pass

    # 等待所有线程结束
    for t in crawl_threads:
        t.join()

    wb = openpyxl.Workbook()
    sheet = wb.active

    for col, head in enumerate(all_data[0].keys()):
        sheet.cell(row=1, column=col+1, value=head)

    for row, sheet_row in enumerate(all_data):
        for col, row_key in enumerate(sheet_row.keys()):
            sheet.cell(row=row+2, column=col+1, value=sheet_row[row_key])
    wb.save(filename="kylinos_compatiable_list.xlsx")

if __name__ == '__main__':
    main()
