# coding=utf-8
import json
import openpyxl
import requests
import re

INLINE_LINK_RE = re.compile(r'\[([^\]]+)\]\(([^)]+)\)')
FOOTNOTE_LINK_TEXT_RE = re.compile(r'\[([^\]]+)\]\[(\d+)\]')
FOOTNOTE_LINK_URL_RE = re.compile(r'\[(\d+)\]:\s+(\S+)')

class Spider:
    def __init__(self):
        self.url = "https://hub.docker.com/api/content/v1/products/search?operating_system=linux&page_size=25&q="
        self.response = []

    def pull_all_data(self):
        payload = {}
        headers = {
            'Accept': 'application/json',
            'Search-Version': 'v3'
            }
        response = json.loads(requests.request("GET", self.url, headers=headers, data=payload).text)
        page_size = response["page_size"]
        numFound = 100
        page_num = numFound // page_size
        for i in range(1, page_num+1):
            url = "https://hub.docker.com/api/content/v1/products/search?operating_system=linux&page_size=25&page=" + str(i)
            payload = {}
            headers = {
                'Accept': 'application/json',
                'Search-Version': 'v3'
            }
            response = json.loads(requests.request("GET", url, headers=headers, data=payload).text)
            next_url = ""
            for item in response['summaries']:
                name = item['name']
                if name.split("/")[0] == item['name']:
                    name = "-".join([i.lower() for i in name.split(" ")])
                    next_url = "https://hub.docker.com/v2/repositories/library/" + name
                else:
                    next_url = "https://hub.docker.com/v2/repositories/" + name.split('/')[0] + "/" + name.split('/')[1] + "/"
                item["dockerhub_url"]  = next_url
                info = self.get_more_information(next_url)
                dictMerged2 = dict(item, **info)
                self.response.append(dictMerged2)

    def get_more_information(self,url):
        response = json.loads(requests.request("GET", url).text)
        links = {}
        if response.get("full_description"):
            links = self.find_md_links(response['full_description'])
        return links

    def find_md_links(self, md):
        """ Return dict of links in markdown """

        links = dict(INLINE_LINK_RE.findall(md))
        footnote_links = dict(FOOTNOTE_LINK_TEXT_RE.findall(md))
        footnote_urls = dict(FOOTNOTE_LINK_URL_RE.findall(md))

        for key, value in footnote_links.items():
            footnote_links[key] = footnote_urls[value]
        links.update(footnote_links)
        urls = {}
        url_list = []
        for i in links.values():
            if "Dockerfile" in i:
                url_list.append(i)
        urls["urls"] = "\n".join(url_list)

        return urls

    def write_to_xlsx(self):
        wb = openpyxl.Workbook()
        sheet = wb.active

        for col, head in enumerate(self.response[0].keys()):
            sheet.cell(row=1, column=col + 1, value=head)

        for row, sheet_row in enumerate(self.response):
            for col, row_key in enumerate(sheet_row.keys()):
                sheet.cell(row=row + 2, column=col + 1, value=str(sheet_row[row_key]))
        wb.save(filename="docker_top_100.xlsx")

if __name__ == '__main__':
    sp = Spider()
    sp.pull_all_data()
    sp.write_to_xlsx()
